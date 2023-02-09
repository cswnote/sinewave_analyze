import os
import sys
import numpy as np
import openpyxl
import csv
import math
import pandas as pd
# import GET_SUMMARY
# import win32com.client
import matplotlib.pyplot as plt
import math
import FILE_MANAGEMENT
import datetime


class tekCsv():
    def __init__(self, **kwargs):
        self.path = kwargs.get('path', 'd:/')
        self.csv_path = kwargs.get('csv_path', self.path + 'tek_csv/')
        self.excel_path = kwargs.get('excel_path', self.path + 'tek_excel/')
        self.test_info_path = kwargs.get('info_path', self.path + 'test information/')
        self.kmon_csv_path = kwargs.get('kmon_path', self.path + 'kmon_csv/')
        self.time_graph = kwargs.get('graph_time', True)
        self.fft_graph = kwargs.get('graph_FFT', True)
        self.filter_factor = kwargs.get('filter_factor', 0.2)
        self.lpf = kwargs.get('LPF', False)
        self.fft = kwargs.get('FFT', True)
        self.time_window_size = kwargs.get('time_window', 10000)
        self.time_window_type = kwargs.get('time_window_type', 'crop')
        self.fft_window_size = kwargs.get('fft_window', 5000)
        self.fft_window_type = kwargs.get('fft_window_type', 'crop')
        self.FFT = kwargs.get('FFT', True)

        self.eval_file = kwargs.get('eval_fila', 'eval_file.xlsx')

        self.record_length = 0


    def get_csv_filelist(self):
        file_list = os.listdir(self.csv_path)
        return [file for file in file_list if file.endswith(".csv")]


    def get_filetype_list(self, filetype, path):
        file_list = os.listdir(path)
        return [file for file in file_list if file.endswith(filetype)]


    def apply_LPF(self, worksheet, data_only=False):
        # print('in process: apply_LPF')
        ws = worksheet

        ws['a17'].value = 'filter_factor'
        ws['b17'].value = self.filter_factor

        if ws['e21'].value is not None:
            number_of_field = 4
        elif ws['d21'].value is not None:
            number_of_field = 3
        elif ws['c21'].value is not None:
            number_of_field = 2
        elif ws['b21'].value is not None:
            number_of_field = 1
        elif ws['b21'].value is None:
            print('Error: data is empty')
            return

        for i in range(number_of_field):
            ws[chr(66 + i + number_of_field) + str(21)].value = 'LPF ' + ws[chr(66 + i) + str(13)].value
            ws[chr(66 + i + number_of_field) + str(13)].value = ws[chr(66 + i) + str(13)].value


            if data_only is False:
                ws[chr(66 + i + number_of_field) + str(20)].value = '= B17'
            else:
                ws[chr(66 + i + number_of_field) + str(20)].value = ws['b17'].value

        for i in range(number_of_field):
            original_column = chr(66 + i)
            target_column = chr(66 + i + number_of_field)

            for j in range(22, self.record_length + 22):
                if data_only is False:
                    if j == 22:
                        ws[target_column + str(j)].value = '= ' + original_column + str(
                            j) + ' * ' + target_column + '20 + ' + original_column + str(
                            j) + ' * (1 - ' + target_column + '20)'
                    else:
                        ws[target_column + str(j)].value = '= ' + original_column + str(
                            j) + '* ' + target_column + '20 + ' + target_column + str(
                            j - 1) + '* (1 - ' + target_column + '20)'

                elif data_only is True:
                    if j == 22:
                        ws[target_column + str(j)].value = float(ws[original_column + str(j)].value) * float(ws[target_column + '20'].value) + float(ws[original_column + str(j)].value) * (1 - float(ws[target_column + '20'].value))
                    else:
                        ws[target_column + str(j)].value = float(ws[original_column + str(j)].value) * float(ws[target_column + '20'].value) + float(ws[target_column + str(j - 1)].value) * (1 - float(ws[target_column + str(20)].value))

        return ws


    def get_fft(self, worksheet):
        # print('in process: get_fft')

        ws = worksheet
        Ts = float(ws['b9'].value)
        Fs = 1/Ts
        t_axis_data = []

        final_t_axis = 'a' + str(self.record_length + 21)
        for rng in ws['a22':final_t_axis]:
            for cell in rng:
                t_axis_data.append(float(cell.value))

        y1_axis_data = []
        final_y1_axis = 'b' + str(self.record_length + 21)
        for rng in ws['b22':final_y1_axis]:
            for cell in rng:
                y1_axis_data.append(float(cell.value))

        n = len(y1_axis_data)  # length of the signal
        k = np.arange(n)
        T = n / Fs  # total measurement time, number of sampling * sampling interval(period of sampling)
        freq = k / T  # two sides frequency range
        freq = freq[range(int(n / 2))]  # one side frequency range
        Y = np.fft.fft(y1_axis_data) / n  # fft computing and normalization
        Y = Y[range(int(n / 2))] * 2 * 0.707

        # fig, ax = plt.subplots(2, 1)
        # ax[0].plot(t_axis_data, y1_axis_data)
        # ax[0].set_xlabel('Time')
        # ax[0].set_ylabel('Amplitude')
        # ax[0].grid(True)
        # ax[1].plot(freq, abs(Y), 'r', linestyle=' ')  # , marker='^')
        # ax[1].set_xlabel('Freq (Hz)')
        # ax[1].set_ylabel('|Y(freq)|')
        # ax[1].vlines(freq, [0], abs(Y))
        # ax[1].grid(True)
        # plt.show()

        return freq, Y


    def cal_fft(self, worksheet, **kwargs):
        # print('in process: cal_fft')

        ws = worksheet
        y_col = kwargs.get('y_col', 'b')
        Ts = float(ws['b9'].value)
        Fs = 1 / Ts
        t_axis_data = []

        final_t_axis = 'a' + str(self.record_length + 21)
        for rng in ws['a22':final_t_axis]:
            for cell in rng:
                t_axis_data.append(float(cell.value))

        y_axis_data = []
        final_y_axis = y_col + str(self.record_length + 21)
        for rng in ws[y_col + '22':final_y_axis]:
            for cell in rng:
                y_axis_data.append(float(cell.value))

        n = len(y_axis_data)  # length of the signal
        k = np.arange(n)
        T = n / Fs  # total measurement time, number of sampling * sampling interval(period of sampling)
        freq = k / T  # two sides frequency range
        freq = freq[range(int(n / 2))]  # one side frequency range
        Y = np.fft.fft(y_axis_data) / n  # fft computing and normalization
        Y = Y[range(int(n / 2))] * 2 / np.sqrt(2)

        return freq, Y


    def draw_chart(self, worksheet, **kwargs):
        # # crop_chart_window 또는 crop_chart_ratio 중 하나만 받아야 함
        # # 'crop_chart_window'의 경우 앞 데이터 crop_chart_window 개수만 그래프 그림
        # # 'crop_chart_ratio'의 경우 '전체 데이터 수/crop_chart_ratio'의 개수만 그래프
        # # 둘다 입력 받을 경우 'crop_chart_window' 우선

        ws = worksheet
        domain = kwargs.get('domain', 'time')
        num_of_chart = kwargs.get('num_of_channel', 1)
        num_of_data_len = kwargs.get('record_length', 10000)
        chart_title = kwargs.get('chart_title', 'chart_name')
        crop_chart_window = kwargs.get('crop_window', sys.maxsize)
        crop_chart_ratio = kwargs.get('crop_ratio', sys.maxsize)

        if domain == 'time':
            chart1 = openpyxl.chart.LineChart()
            chart2 = openpyxl.chart.LineChart()

            if num_of_data_len > 1000000 - 21:
                num_of_data_len = 1000000 - 21
            if crop_chart_ratio != sys.maxsize and crop_chart_window == sys.maxsize:
                num_of_data_len = int(num_of_data_len / crop_chart_ratio)
            elif crop_chart_ratio == sys.maxsize and crop_chart_window != sys.maxsize:
                num_of_data_len = crop_chart_window
            elif crop_chart_ratio != sys.maxsize and crop_chart_window != sys.maxsize:
                num_of_data_len = crop_chart_window

            v_col = 1
            i_col = 1
            for i in range(9, 1, -1):
                if ws.cell(13, i).value == 'V':
                    v_col = i
                elif ws.cell(13, i).value == 'A':
                    i_col = i

                if v_col != 1 and i_col != 1:
                    break

            for i in range(9, 1, -1):
                if v_col == 1 and i_col == 1:
                    break

                if ws.cell(13, i).value == 'V':
                    v_min = sys.maxsize
                    v_max = -sys.maxsize

                    for j in range(num_of_data_len):
                        if float(ws.cell(22 + j, i).value) >= v_max:
                            v_max = float(ws.cell(22 + j, i).value)
                        if float(ws.cell(22 + j, i).value) <= v_min:
                            v_min = float(ws.cell(22 + j, i).value)

                    v_max, v_min = self.get_y_axis_min_max(v_max, v_min)

                    chart1.title = chart_title
                    chart1.style = 10
                    chart1.x_axis.title = "time"
                    chart1.y_axis.scaling.min = v_min
                    chart1.y_axis.scaling.max = v_max
                    data1 = openpyxl.chart.Reference(ws, min_col=v_col, max_col=v_col,
                                                     min_row=21, max_row=num_of_data_len + 21)
                    cats = openpyxl.chart.Reference(ws, min_col=1, min_row=22, max_row=num_of_data_len + 21)  # 축 설정
                    chart1.add_data(data1, titles_from_data=True)
                    chart1.set_categories(cats)
                    chart1.series[0].graphicalProperties.line.width = 0
                    chart1.y_axis.majorGridlines = None

                    v_col = 1

                elif ws.cell(13, i).value == 'A':
                    i_min = sys.maxsize
                    i_max = -sys.maxsize

                    for j in range(num_of_data_len):
                        if float(ws.cell(22 + j, i).value) >= i_max:
                            i_max = float(ws.cell(22 + j, i).value)
                        elif float(ws.cell(22 + j, i).value) <= i_min:
                            i_min = float(ws.cell(22 + j, i).value)

                    i_max, i_min = self.get_y_axis_min_max(i_max, i_min)

                    chart2.title = chart_title
                    chart2.style = 10
                    chart2.y_axis.scaling.min = i_min
                    chart2.y_axis.scaling.max = i_max
                    data2 = openpyxl.chart.Reference(ws, min_col=i_col, max_col=i_col,
                                                     min_row=21, max_row=num_of_data_len + 21)
                    cats = openpyxl.chart.Reference(ws, min_col=1, min_row=22, max_row=num_of_data_len + 21)  # 축 설정
                    chart2.add_data(data2, titles_from_data=True)
                    chart2.set_categories(cats)
                    chart2.series[0].graphicalProperties.line.width = 0
                    chart2.y_axis.majorGridlines = None

                    chart2.y_axis.axId = 200
                    chart2.y_axis.crosses = 'max'
                    i_col = 1

            chart1 += chart2

            chart_location = chr(num_of_chart * 2 + 2 + ord('a'))

            ws.add_chart(chart1, chart_location + '21')
            chart1.width = 30
            chart1.height = 18

        elif domain == 'frequency':
            for i in range(500002, 0, -1):
                if type(ws.cell(i, 1).value) is int or type(ws.cell(i, 1).value) is float:
                # if ws.cell(i, 1).value == 'None':
                    num_of_data_len = i - 1
                    break

            if crop_chart_ratio != sys.maxsize and crop_chart_window == sys.maxsize:
                num_of_data_len = int(num_of_data_len / crop_chart_ratio)
            elif crop_chart_ratio == sys.maxsize and crop_chart_window != sys.maxsize:
                num_of_data_len = crop_chart_window
            elif crop_chart_ratio != sys.maxsize and crop_chart_window != sys.maxsize:
                num_of_data_len = crop_chart_window

            chart1 = openpyxl.chart.BarChart()
            chart1.title = chart_title
            chart1.style = 20
            chart1.x_axis.title = "frequency"

            data1 = openpyxl.chart.Reference(ws, min_col=3, max_col=3, min_row=1,
                                             max_row=num_of_data_len + 1)
            cats = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=num_of_data_len + 1)  # 축 설정

            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats)

            for i in range(num_of_chart):
                chart1.series[i].graphicalProperties.line.width = 0

            chart_location = chr(num_of_chart + 2 + ord('a'))

            ws.add_chart(chart1, 'e6')
            chart1.width = 30
            chart1.height = 18


    def get_VI_delay(self, frequency, data_len, worksheet, wieghting_num=69):
        # print('in process: get_VI_delay')
        ws = worksheet

        if frequency != 0:
            T = 1/frequency
        else:
            return 0, 0

        v_times = []
        v_freq = []
        v_location = []
        i_times = []
        i_location = []
        vi_diff = []
        vi_angle = []
        vi_coefficient = []
        v_mean = 0.
        i_mean = 0.

        vc_in_window = 0
        ic_in_window = 0

        # # rising or falling
        sample_times = int(abs(T/(float(ws.cell(23, 1).value) - float(ws.cell(24, 1).value))) * 0.8)
        # # rising and falling
        # sample_times = int(abs(T/(float(ws.cell(23, 1).value) - float(ws.cell(24, 1).value))) * 0.8) / 2

        v_trigger = False
        i_trigger = False

        for i in range(9, 1, -1):

            if ws.cell(13, i).value == 'V' and not v_trigger:
                v_trigger = True
                ws['g2'].value = 'V rising time'
                ws['g1'].value = 'V freq[MHz]'
                max_value = -(sys.maxsize+1)
                min_value = sys.maxsize

                for j in range(22, data_len + 22):
                    if max_value < float(ws.cell(j, i).value):
                        max_value = float(ws.cell(j, i).value)
                    if min_value > float(ws.cell(j, i).value):
                        min_value = float(ws.cell(j, i).value)

                vc_in_window = (max_value + min_value)/2

                # for j in range(23, data_len + 22):
                #     if float(ws.cell(j - 1, i).value) <= vc_in_window <= float(ws.cell(j, i).value):
                #         if abs(float(ws.cell(j - 1, i).value) - vc_in_window) > abs(float(ws.cell(j, i).value) - vc_in_window):
                #             v_times.append(float(ws.cell(j, 1).value))
                #             j = j + sample_times
                #         else:
                #             v_times.append(float(ws.cell(j - 1, 1).value))
                #             j = j + sample_times
                j = 23
                while j < (data_len + 22):
                    # rising
                    if float(ws.cell(j - 1, i).value) <= vc_in_window <= float(ws.cell(j, i).value):
                        flag = True
                        for k in range(wieghting_num):
                            try:
                                if float(ws.cell(j - 1, i).value) > float(ws.cell(j + k, i).value):
                                    flag = False
                                    break
                            except:
                                pass
                        if flag:
                            if abs(float(ws.cell(j - 1, i).value) - vc_in_window) >= abs(float(ws.cell(j, i).value) - vc_in_window):
                                v_times.append(float(ws.cell(j, 1).value))
                                v_location.append(j)
                                j = j + sample_times
                            else:
                                v_times.append(float(ws.cell(j - 1, 1).value))
                                v_location.append(j - 1)
                                j = j + sample_times
                    # falling
                    # elif float(ws.cell(j - 1, i).value) >= vc_in_window >= float(ws.cell(j, i).value):
                    #     if abs(float(ws.cell(j - 1, i).value) - vc_in_window) > abs(float(ws.cell(j, i).value) - vc_in_window):
                    #         v_times.append(float(ws.cell(j, 1).value))
                    #         j = j + sample_times
                    #     else:
                    #         v_times.append(float(ws.cell(j - 1, 1).value))
                    #         j = j + sample_times

                    j = j + 1


            elif ws.cell(13, i).value == 'A' and not i_trigger:
                i_trigger = True
                ws['g3'].value = 'I rising time'
                max_value = -(sys.maxsize+1)
                min_value = sys.maxsize

                for j in range(22, data_len):
                    if max_value < float(ws.cell(j, i).value):
                        max_value = float(ws.cell(j, i).value)
                    if min_value > float(ws.cell(j, i).value):
                        min_value = float(ws.cell(j, i).value)

                ic_in_window = round((max_value + min_value) / 2, 8)

                # for j in range(23, data_len + 22):
                #     if float(ws.cell(j - 1, i).value) <= vc_in_window <= float(ws.cell(j, i).value):
                #         if abs(float(ws.cell(j - 1, i).value) - vc_in_window) > abs(float(ws.cell(j, i).value) - vc_in_window):
                #             v_times.append(float(ws.cell(j, 1).value))
                #             j = j + sample_times
                #         else:
                #             v_times.append(float(ws.cell(j - 1, 1).value))
                #             j = j + sample_times
                j = 23
                while j < (data_len + 22):
                    # rising
                    if float(ws.cell(j - 1, i).value) <= ic_in_window <= float(ws.cell(j, i).value):
                        flag = True
                        for k in range(wieghting_num):
                            try:
                                if float(ws.cell(j - 1, i).value) > float(ws.cell(j + k, i).value):
                                    flag = False
                                    break
                            except:
                                pass
                        if flag:
                            if abs(float(ws.cell(j - 1, i).value) - ic_in_window) >= abs(float(ws.cell(j, i).value) - ic_in_window):
                                i_times.append(float(ws.cell(j, 1).value))
                                i_location.append(j)
                                j = j + sample_times
                            else:
                                i_times.append(float(ws.cell(j - 1, 1).value))
                                i_location.append(j - 1)
                                j = j + sample_times
                    # falling
                    # elif float(ws.cell(j - 1, i).value) >= vc_in_window >= float(ws.cell(j, i).value):
                    #     if abs(float(ws.cell(j - 1, i).value) - vc_in_window) > abs(float(ws.cell(j, i).value) - vc_in_window):
                    #         v_times.append(float(ws.cell(j, 1).value))
                    #         j = j + sample_times
                    #     else:
                    #         v_times.append(float(ws.cell(j - 1, 1).value))
                    #         j = j + sample_times

                    j = j + 1

            elif v_trigger and i_trigger:
                break


        ws.cell(4, 7, value='difference')
        ws.cell(5, 7, value='angle')
        ws.cell(6, 7, value='real power Coefficient')
        ws.cell(7, 7, value='ave.RP Co')
        ws.cell(1, 5, value='Vmean')
        ws.cell(3, 5, value='Imean')

        if len(v_times) > len(i_times):
            times = len(i_times)
            f_gap = abs(v_times[0] - i_times[0])
            e_gap = abs(v_times[len(v_times) - 1] - i_times[len(i_times) - 1])
            if f_gap > e_gap:
                del v_times[0]
                del v_location[0]
            else:
                del v_times[len(v_times) - 1]
                del v_location[len(v_location) - 1]
        elif len(v_times) < len(i_times):
            times = len(v_times)
            f_gap = abs(v_times[0] - i_times[0])
            e_gap = abs(v_times[len(v_times) - 1] - i_times[len(i_times) - 1])
            if f_gap > e_gap:
                del i_times[0]
                del i_location[0]
            else:
                del i_times[len(i_times) - 1]
                del i_location[len(i_times) - 1]
        else:
            if v_times[0] - i_times[0] > T/2:
                del i_times[0]
                del v_times[len(v_times) - 1]
            elif i_times[0] - v_times[0] > T/2:
                del v_times[0]
                del i_times[len(i_times) - 1]
            times = len(v_times)

        for i in range(len(i_times)):
            ws.cell(3, 8 + i, value=i_times[i])
        for i in range(len(v_times)):
            ws.cell(2, 8 + i, value=v_times[i])

        for j in range(len(v_times) - 1):
            v_freq.append(1 / (v_times[j + 1] - v_times[j]) * 10 ** (-6))
            ws.cell(1, 8 + j, value=v_freq[j])
        ws.cell(1, 8 + len(v_freq) + 1).value = (sum(v_freq, 0.0) / len(v_freq))


        for i in range(times):
            vi_diff.append(v_times[i] - i_times[i])
            vi_angle.append(vi_diff[i]/T * 360)
            ws.cell(4, 8 + i, value=v_times[i] - i_times[i])

            # if vi_angle[i] < 0:
            #     vi_angle[i] = 360 + vi_angle[i]

            ws.cell(5, 8 + i, value=vi_angle[i])

            vi_coefficient.append(math.cos(math.radians(vi_angle[i])))

            ws.cell(6, 8 + i, value=vi_coefficient[i])

        ws.cell(5, 8 + len(vi_angle)).value = sum(vi_angle, 0.0)/len(vi_angle)

        real_power_mean = np.mean(np.array(vi_coefficient))
        ws.cell(7, 8, value=real_power_mean)

        v_trigger = False
        i_trigger = False
        for i in range(9, 1, -1):
            if ws.cell(13, i).value == 'V' and not v_trigger:
                for j in range(v_location[0], v_location[-1] + 1):
                    v_mean += float(ws.cell(j, i).value)
                v_mean = v_mean / (v_location[-1] - v_location[0] + 1)
                v_trigger = True
            elif ws.cell(13, i).value == 'A' and not i_trigger:
                for j in range(i_location[0], i_location[-1] + 1):
                    i_mean += float(ws.cell(j, i).value)
                i_mean = i_mean / (i_location[-1] - i_location[0] + 1)
                i_trigger = True
            elif v_trigger and i_trigger:
                break
        # ws.cell(11, 8, value=v_mean)
        ws.cell(1, 6, value=v_mean)
        # ws.cell(12, 8, value=i_mean)
        ws.cell(3, 6, value=i_mean)

        return v_location, i_location


    def get_rms(self, v_row_nums, i_row_nums, data_len, worksheet):
        # print('in process: get_rms')
        ws = worksheet

        v_data = []
        i_data = []
        v_rms = []
        i_rms = []
        rp_rms = []

        v_trigger = 1
        i_trigger = 1

        for i in range(9, 1, -1):
            if ws.cell(13, i).value == 'V' and v_trigger == 1:
                v_trigger = 0
                # per one period
                for j in range(len(v_row_nums) - 1):
                    for k in range(v_row_nums[j], v_row_nums[j + 1]):
                        v_data.append(float(ws.cell(k, i).value))
                    v_np = np.array(v_data)
                    v_rms.append(np.sqrt(np.mean(v_np**2)))
                    v_data = []
                # # all window
                # for j in range(22, data_len + 22):
                #     v_data.append(float(ws.cell(j, i).value))
                # v_np = np.array(v_data)
                # v_rms = np.sqrt(np.mean(v_np**2))

                ws['g8'] = 'Vrms_' + ws.cell(21, i).value

            elif ws.cell(13, i).value == 'A' and i_trigger == 1:
                i_trigger = 0
                for j in range(len(i_row_nums) - 1):
                    for k in range(i_row_nums[j], i_row_nums[j + 1]):
                        i_data.append(float(ws.cell(k, i).value))
                    i_np = np.array(i_data)
                    i_rms.append(np.sqrt(np.mean(i_np**2)))
                    i_data = []
                # # all window
                # for j in range(22, data_len + 22):
                #     i_data.append(float(ws.cell(j, i).value))
                # i_np = np.array(i_data)
                # i_rms = np.sqrt(np.mean(i_np**2))

                ws['g9'] = 'Irms_' + ws.cell(21, i).value

        for i in range(len(v_rms)):
            ws.cell(8, 8 + i, value=v_rms[i])
        for i in range(len(i_rms)):
            ws.cell(9, 8 + i, value=i_rms[i])

        ws['g10'] = 'Real Power'

        if len(v_rms) == len(i_rms):
            for i in range(len(v_rms)):
                Vrms = ws.cell(8, 8 + i).value
                Irms = ws.cell(9, 8 + i).value
                RPf = ws.cell(6, 8 + 1).value
                rp_rms.append(Vrms * Irms * RPf)
                ws.cell(10, 8 + i, value=Vrms * Irms * RPf)

            ws.cell(7, 8 + len(v_rms) + 1).value = 'average'
            ws.cell(8, 8 + len(v_rms) + 1).value = (sum(v_rms, 0)/len(v_rms))
            ws.cell(9, 8 + len(i_rms) + 1).value = (sum(i_rms, 0)/len(i_rms))
            ws.cell(10, 8 + len(rp_rms) + 1).value = (sum(rp_rms, 0) / len(rp_rms))


    def get_y_axis_min_max(self, max, min):
        max_scaled = 0
        max_scaled = 0

        if max == float('inf'):
            max = 20
        if min == float('-inf'):
            min = -20

        digits = len(str(float(abs(max))).split('.')[0])
        if max > 0:

            max_scaled = math.ceil(max / (10 ** (digits - 1))) * 10 ** (digits - 1)
        else:
            max_scaled = math.ceil(max / (10 ** (digits - 1))) * 10 ** (digits - 1) * 1

        digits = len(str(float(abs(min))).split('.')[0])

        if min > 0:
            min_scaled = math.ceil(min / (10 ** (digits - 1))) * 10 ** (digits - 1)
        else:
            min_scaled = math.floor(min / (10 ** (digits - 1))) * 10 ** (digits - 1)

        if abs(max_scaled) > abs(min_scaled):
            min_scaled = -max_scaled
        else:
            max_scaled = -min_scaled

        if max_scaled == min_scaled or max_scaled < min_scaled:
            return max, min
        else:
            return max_scaled, min_scaled


    def get_pulse_width(self, wb, **kwargs):
        ch_num1 = kwargs.get('ch_num1', 'CH3')
        ch_num2 = kwargs.get('ch_num2', 'CH4')

        ws = wb['FFT_' + ch_num1]
        ch_num1_freq = ws['f1'].value
        ch_num1_bias = (ws['c2'].value) / np.sqrt(2)
        ws = wb['FFT_' + ch_num2]
        ch_num2_freq = ws['f1'].value
        ch_num2_bias = (ws['c2'].value) / np.sqrt(2)
        ws = wb[wb.get_sheet_names()[0]]

        print(ws)


    def get_test_file_list(self, df):
        for i in range(len(df)):
            print(df[df[i]])


    def combine_infofiles(self):
        file_list = self.get_filetype_list('xlsx', self.test_info_path)
        file_list.sort()
        file_list = [file for file in file_list if file[:10] == 'info_test_']

        if 'info_test_all.xlsx' in file_list:
            df = pd.read_excel(self.test_info_path + 'info_test_all.xlsx')
            file_list.remove('info_test_all.xlsx')
        else:
            if not len(file_list):
                print('''this directory has not none 'info_test' file''')
                return
            else:
                df = pd.read_excel(self.test_info_path + file_list[0])
                file_list.remove(file_list[0])

        try:
            for idx, file in enumerate(file_list):
                df_add = pd.read_excel(self.test_info_path + file)
                df = pd.concat([df, df_add], ignore_index=True)
        except:
            print('fail to merge info data')

        df.set_index(keys=df.columns[0], drop=True, inplace=True)
        df.to_excel(self.test_info_path + 'info_test_all.xlsx', index_label=df.columns[0])

        print('end add_info_file')

    def csv_to_excel(self, lpf_factor=0.5, lpf=False):
        # csv_list = self.get_csv_filelist()
        # csv_list.sort()

        csv_list = self.get_filetype_list('csv', self.csv_path)
        csv_list.sort()

        for idx, csv_file in enumerate(csv_list):
            previous_time = datetime.datetime.now()
            print('in process: ', idx + 1, '/', len(csv_list), '    ', csv_file, end='    ')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = csv_file.split('.')[0]

            COL_SEPARATOR = ","
            with open(self.csv_path + csv_file) as file:
                reader = csv.reader(file)
                for r, row in enumerate(reader):
                    for c, col in enumerate(row):
                        for idx, val in enumerate(col.split(COL_SEPARATOR)):
                            cell = ws.cell(row=r+1, column=c+1)
                            cell.value = val
                            cell.data_type = 'General'

                # # 아래는 ch2, ch3 할 때 살릴 것, ch3 ch4 에서는 주석 처리
                # # 5, 4 삭제는 ch1, ch2 남음
                # ws.delete_cols(8)
                # ws.delete_cols(7)
                # ws.delete_cols(6)
                # ws.delete_cols(5)
                # ws.delete_cols(4)

                self.record_length = int(float(ws['b10'].value))

                # # can use cal_fft() when data_only is True
                if self.lpf:
                    ws = self.apply_LPF(ws, data_only=True)

                num_channel = 0
                for i in range(9, 1, -1):
                    # # get number of channel
                    if ws.cell(13, i).value is not None:
                            num_channel = num_channel + 1

                if self.time_graph:
                    chart_name = ws.title
                    if self.time_window_type == 'crop':
                        self.draw_chart(ws, num_of_channel=num_channel, record_length=self.record_length, domain='time',
                                        chart_title=chart_name, crop_window=self.time_window_size)
                    elif self.time_window_type == 'ratio':
                        self.draw_chart(ws, num_of_channel=num_channel, record_length=self.record_length, domain='time',
                                        chart_title=chart_name, crop_ratio=self.time_window_size)


                if self.FFT:
                    for i in range(num_channel):
                        fft_column = chr(i + 98)
                        freq, amplitude = self.cal_fft(ws, y_col=fft_column)

                        ws_fft = wb.create_sheet(title='FFT_' + ws[fft_column + '21'].value)

                        ws_fft['a1'].value = 'freq'
                        ws_fft['b1'].value = 'Y_complex'
                        ws_fft['c1'].value = 'Y_absolute'

                        ws_fft['e1'].value = 'F @ max Y'
                        ws_fft['e2'].value = 'max Y abs'

                        for j in range(len(freq)):
                            ws_fft.cell(j + 2, 1, value=freq[j])
                            ws_fft.cell(j + 2, 2, value='= complex(' + str(amplitude[j].real) + ', ' + str(amplitude[j].imag) + ', "j")')
                            ws_fft.cell(j + 2, 3, value=abs(amplitude[j]))

                        if self.fft_graph:
                            if self.fft_window_type == 'crop':
                                self.draw_chart(ws_fft, domain='frequency', crop_window=self.fft_window_size)
                            elif self.fft_window_type:
                                self.draw_chart(ws_fft, domain='frequency', crop_ratio=self.fft_window_size)

                        max_amplitude = 0
                        max_freq = 0
                        for j in range(1, len(freq)):
                            if float(ws_fft.cell(j + 2, 3).value) > max_amplitude:
                                max_amplitude = float(ws_fft.cell(j + 2, 3).value)
                                max_freq = float(ws_fft.cell(j + 2, 1).value)

                        ws_fft['f1'].value = max_freq
                        ws_fft['f2'].value = max_amplitude

                        wb.save(self.excel_path + csv_file.split('.csv')[0] + '.xlsx')

                    v_row_num, i_row_num = self.get_VI_delay(max_freq, self.record_length, ws, wieghting_num=100)

                    if v_row_num != 0:
                        self.get_rms(v_row_num, i_row_num, self.record_length, ws)

            wb.save(self.excel_path + csv_file.split('.csv')[0] + '.xlsx')
            wb.close()

            now_time = datetime.datetime.now()
            remain_time = (now_time - previous_time).seconds * (len(csv_list) - idx + 1)
            remain_time = remain_time // 60
            print("remain time is {}minutes".format(remain_time))


    def file_name_change(self, sheet):
        df_name = pd.read_excel(self.path + self.eval_file, sheet_name=sheet)

        info_files = df_name.iloc[:, 0].tolist()
        # info_files = [file for file in info_files if file[:10] == 'info_test_' and not('all' in file)]
        test_files = os.listdir(self.excel_path)
        test_files = [file for file in test_files if file.endswith('xlsx') and file[:3] == 'tek']
        test_files.sort()

        for i in range(len(df_name)):
            df_test_info = pd.read_excel(self.test_info_path + df_name.at[i, 'filename'] + '.xlsx')
            start = int(df_test_info.at[0, 'filename'][3:])
            end = int(df_test_info.at[len(df_test_info) - 1, 'filename'][3:])
            for file in test_files:
                if start <= int(file.split(' ')[0][3:].split('.')[0]) <= end:
                    extension = file.split('.')[1]
                    file = file.split('.')[0].split(' ')[0]
                    scr = self.excel_path + file + '.' + extension

                    idx = df_test_info.index[df_test_info['filename'] == file].tolist()[0]

                    for j in range(1, len(df_name.columns)):
                        if 'field' in df_name.columns[j]:
                            if type(df_name.at[i, df_name.columns[j]]) is str:
                                column = df_name.at[i, df_name.columns[j]]
                                # if len(column.split(' ')) >= 2:
                                #     if column.lower() == 'pwm':
                                #         column = 'CP Pwm Set'
                                #         file = file + ' ' + str(column).split(' ')[1] + str(
                                #             df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])
                                # else:
                                #     if column.lower() == 'pwm':
                                #         column = 'CP Pwm Set'
                                #         file = file + ' ' + str(column).split(' ')[1] + str(
                                #             df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])
                                if 'pwm' in column.lower():
                                    column = 'CP Pwm Set'
                                    file = file + ' ' + str(column).split(' ')[1] + str(
                                        df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])
                                elif 'volt' in column.lower():
                                    column = 'RF Volt Set'
                                    file = file + ' ' + str(column).split(' ')[1] + str(
                                        df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])
                                elif 'curr' in column.lower():
                                    column = 'RF Curr Set'
                                    file = file + ' ' + str(column).split(' ')[1] + str(
                                        df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])

                            elif type(df_name.at[i, df_name.columns[j]]) is float:
                                if not math.isnan(df_name.at[i, df_name.columns[j]]):
                                    column = str(df_name.at[i, df_name.columns[j]])
                                    # file = file + ' ' + str(column) + ' ' + str(df_test_info.at[idx, column])
                                    file = file + ' ' + str(column).split(' ')[1] + \
                                           str(df_test_info.at[idx, column + ' ' + df_name.at[i, 'channel']])

                        else:
                            if 'ohm' == df_name.columns[j].lower():
                                file = file + ' ' + (str(df_name.at[i, df_name.columns[j]]) + 'ohm').replace(' ', '')
                            else:
                                file = file + ' ' + str(df_name.at[i, df_name.columns[j]]).replace(' ', '')
                            # file = file + str(df_name.iat[i, j]) # 위와 동일

                    dst = self.excel_path + file + '.' + extension
                    os.rename(scr, dst)