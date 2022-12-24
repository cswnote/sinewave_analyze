import openpyxl
import os
import numpy as np
import pandas as pd
import re
import math
import gc
import platform
# from win32com.client import Dispatch



class Get_Summary():
    def __init__(self, path, eval_file):
        super().__init__()

        self.path = path
        self.tek_csv_path = path + 'tek_csv/'
        self.tek_excel_path = path + 'tek_excel/'
        self.kmon_csv_path = path + 'kmon_csv/'
        self.test_info_path = path + 'test information/'
        self.eval_file = eval_file

        self.measure_value = {'filename': []}
        self.lost_files = []

    def get_summary(self):
        path = self.tek_excel_path # excel path로 변경
        excel_list = os.listdir(self.tek_excel_path)
        excel_list = [file for file in excel_list if file[:3] == 'tek' and file.endswith('.xlsx')]
        excel_list.sort()

        summary_wb = openpyxl.Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = 'summary'

        # summary_ws('a1').value = excel_file.spilt('.xlsx')[0]

        # summary_ws['a1'].value = 'filename'
        # summary_ws['b1'].value = 'V Frequency[MHz]'
        # summary_ws['c1'].value = 'Delay(degree)'
        # summary_ws['d1'].value = 'Ave. RP Coff'
        # summary_ws['e1'].value = 'Vrms'
        # summary_ws['f1'].value = 'Irms'
        # summary_ws['g1'].value = 'Real P[W]'
        # summary_ws['h1'].value = 'V dc'
        # summary_ws['i1'].value = 'I dc'
        # summary_ws['j1'].value = 'FFT V freq[MHz]'
        # summary_ws['k1'].value = 'FFT V rms'
        # summary_ws['l1'].value = 'FFT V dc abs'
        # summary_ws['m1'].value = 'FFT I freq[MHz]'
        # summary_ws['n1'].value = 'FFT I rms'
        # summary_ws['o1'].value = 'FFT I dc abs'

        for idx, excel_file in enumerate(excel_list):
            print('in summary process: ', idx + 1, '/', len(excel_list), '    ', excel_file)
            wb = openpyxl.load_workbook(path + excel_file)
            try:
                ws = wb[excel_file.split(' ')[0]]
            except:
                ws = wb[excel_file.split('.')[0]]
            # summary_ws.cell(idx + 2, 1).value = excel_file.split('.xlsx')[0]

            # file_space = len(excel_file[:excel_file.index('Ch') + 3].split(' '))
            file_space = 7

            summary_name = False
            if not summary_name:
                summary_ws['a1'].value = 'filename'
                summary_ws['b1'].value = 'Board'
                summary_ws.cell(1, 3).value = 'ohm'
                summary_ws.cell(1, 4).value = 'Ch'
                summary_ws.cell(1, 5).value = 'PWM'
                summary_ws.cell(1, 6).value = 'Volt'
                summary_ws.cell(1, 7).value = 'Curr'

                ohm = 3
                ch = 4
                PWM = 5
                Volt = 6
                Curr = 7
                fields = ['ohm', 'ch', 'PWM', 'Volt', 'Curr']

                summary_ws.cell(1, file_space + 1).value = 'V Frequency[MHz]'
                summary_ws.cell(1, file_space + 2).value = 'Delay(degree)'
                summary_ws.cell(1, file_space + 3).value = 'Ave. RP Coff'
                summary_ws.cell(1, file_space + 4).value = 'Vpeak[V]'
                summary_ws.cell(1, file_space + 5).value = 'Irms[mA]'
                summary_ws.cell(1, file_space + 6).value = 'Real P[W]'
                summary_ws.cell(1, file_space + 7).value = 'Vmean'
                summary_ws.cell(1, file_space + 8).value = 'Imean[mA]'
                summary_ws.cell(1, file_space + 9).value = 'FFT V freq[MHz]'
                summary_ws.cell(1, file_space + 10).value = 'FFT V rms'
                summary_ws.cell(1, file_space + 11).value = 'FFT V dc abs'
                summary_ws.cell(1, file_space + 12).value = 'FFT I freq[MHz]'
                summary_ws.cell(1, file_space + 13).value = 'FFT I rms[mA]'
                summary_ws.cell(1, file_space + 14).value = 'FFT I dc abs[mA]'


                # for i, name in enumerate(excel_file[:excel_file.index('Ch') + 3].split(' ')):
                #     if i > 1:
                #         if name[:2].lower() == 'ch':
                #             summary_ws.cell(1, i + 1).value = 'ch'
                #             ch = i + 1
                #         elif name[-3:].lower() == 'ohm':
                #             summary_ws.cell(1, i + 1).value = 'ohm'
                #             ohm = i + 1
                #         else:
                #             temp = re.findall('[0-9]+', name)
                #             summary_ws.cell(1, i + 1).value = name[:-len(temp[0])]
                #             fields.append(name[:-len(temp[0])])




                summary_name = True

            # for i, name in enumerate(excel_file.split('.xlsx')[0].split(' ')):
            #     if name[:2].lower() == 'ch':
            #         summary_ws.cell(idx + 2, i + 1).value = 'ch'
            #         ch = i + 1
            #     elif name[-3:].lower() == 'ohm':
            #         summary_ws.cell(idx + 2, i + 1).value = 'ohm'
            #         ohm = i + 1

            summary_ws.cell(idx + 2, 1).value = excel_file.split('.xlsx')[0].split(' ')[0]
            summary_ws.cell(idx + 2, 2).value = excel_file.split('.xlsx')[0].split(' ')[1]

            summary_ws.cell(idx + 2, ohm).value = excel_file.split('ohm')[0].split(' ')[-1]
            summary_ws.cell(idx + 2, ch).value = 'Ch' + str(excel_file.split('Ch')[1].split(' ')[0])
            try:
                summary_ws.cell(idx + 2, PWM).value = excel_file.split('.xlsx')[0].split('Pwm')[1].split(' ')[0]
            except:
                summary_ws.cell(idx + 2, PWM).value = '-'
            try:
                summary_ws.cell(idx + 2, Volt).value = excel_file.split('.xlsx')[0].split('Volt')[1].split(' ')[0]
            except:
                summary_ws.cell(idx + 2, Volt).value = '-'
            try:
                summary_ws.cell(idx + 2, Curr).value = excel_file.split('.xlsx')[0].split('Curr')[1].split(' ')[0]
            except:
                summary_ws.cell(idx + 2, Curr).value = '-'



            # summary_ws.cell(idx + 2, ch).value = excel_file.split('.xlsx')[0].split(' ')[ch - 1]
            # if 'ohm' in excel_file:
            #     summary_ws.cell(idx + 2, ohm).value =
            # for i in range(len(fields)):
            #     if ch > ohm:
            #         item = ch
            #     else:
            #         item = ohm
            #     if fields[i].lower() == 'pwm':
            #         # # 확인 필요
            #         summary_ws.cell(idx + 2, item + i + 1).value = int(excel_file.split('.xlsx')[0].split(' ')[item + i].split(fields[i])[-1])
            #     else:
            #         summary_ws.cell(idx + 2, item + i + 1).value = excel_file.split('.xlsx')[0].split(' ')[item + i]


                # if i == file_space - 2:
                #     summary_ws.cell(idx + 2, i + 1).value = float(
                #         (excel_file.split('.xlsx')[0].split(' ')[i]).split('ohm')[0])
                # if i == file_space - 1:
                #     summary_ws.cell(idx + 2, i + 1).value = int((excel_file.split('.xlsx')[0].split(' ')[i]).split('PWM')[1])

            # # Delay(degree)
            for i in range(100, 8, -1):
                if ws.cell(5, i).value is not None:
                    summary_ws.cell(idx + 2, file_space + 2).value = ws.cell(5, i).value
                    break

            # # V freq
            for i in range(100, 8, -1):
                if ws.cell(1, i).value is not None:
                    summary_ws.cell(idx + 2, file_space + 1).value = ws.cell(1, i).value
                    break

            # # Ave. RP Coff
            summary_ws.cell(idx + 2, file_space + 3).value = ws.cell(7, 8).value

            # # Vpeak
            for i in range(100, 8, -1):
                if ws.cell(8, i).value is not None:
                    summary_ws.cell(idx + 2, file_space + 4).value = ws.cell(8, i).value * np.sqrt(2)
                    break

            # # Irms
            for i in range(100, 8, -1):
                if ws.cell(9, i).value is not None:
                    summary_ws.cell(idx + 2, file_space + 5).value = ws.cell(9, i).value * 1000
                    break

            # # real power
            for i in range(100, 8 , -1):
                if ws.cell(10, i).value is not None:
                    summary_ws.cell(idx + 2, file_space + 6).value = ws.cell(10, i).value
                    break

            # # Vdc
            summary_ws.cell(idx + 2, file_space + 7).value = ws['f1'].value

            # # Idc
            a = ws['f3'].value
            try:
                summary_ws.cell(idx + 2, file_space + 8).value = ws['f3'].value * 1000
            except:
                summary_ws.cell(idx + 2, file_space + 8).value = ws['f3'].value


            # # get FFT info

            v_flag = False
            i_flag = False
            for i in range(9, 1, -1):
                if ws.cell(13, i).value == 'V' and not v_flag:
                    ws_fft = wb['FFT_' + ws.cell(21, i).value]
                    try:
                        summary_ws.cell(idx + 2, file_space + 9).value = float(ws_fft['f1'].value / 10**6)
                    except:
                        summary_ws.cell(idx + 2, file_space + 9).value = ws_fft['f1'].value
                    summary_ws.cell(idx + 2, file_space + 10).value = ws_fft['f2'].value
                    try:
                        summary_ws.cell(idx + 2, file_space + 11).value = ws_fft['c2'].value / np.sqrt(2)
                    except:
                        summary_ws.cell(idx + 2, file_space + 11).value = ws_fft['c2'].value
                    v_flag = True
                elif ws.cell(13, i).value == 'A' and not i_flag:
                    ws_fft = wb['FFT_' + ws.cell(21, i).value]
                    try:
                        summary_ws.cell(idx + 2, file_space + 12).value = float(ws_fft['f1'].value / 10**6)
                    except:
                        summary_ws.cell(idx + 2, file_space + 12).value = ws_fft['f1'].value
                    summary_ws.cell(idx + 2, file_space + 13).value = ws_fft['f2'].value * 1000
                    try:
                        summary_ws.cell(idx + 2, file_space + 14).value = ws_fft['c2'].value / np.sqrt(2) * 1000
                    except:
                        summary_ws.cell(idx + 2, file_space + 14).value = ws_fft['c2'].value
                elif v_flag and i_flag:
                    break

            wb.close()

        summary_wb.save(path + 'summary.xlsx')
        summary_wb.close()

    # def copy_paste_graph(self, **kwargs):
    #     path = kwargs.get('path', os.getcwd() + '\\')
    #     file_orders = kwargs.get('file_list', 'Sheet1')
    #     summary = kwargs.get('summary_file_name', 'summary')
    #
    #     file_list = os.listdir(path)
    #     file_list = [file for file in file_list if file.endswith(".xlsx")]
    #
    #     for list in file_list:
    #         if list[0:3] != 'tek' or list[-4:] != 'xlsx':
    #             file_list.remove(list)
    #
    #     summary_wb = openpyxl.load_workbook(path + summary + '.xlsx')
    #     order_ws = summary_wb[file_orders]
    #
    #     graph_head = []
    #
    #     for i in range(1, 16385):
    #         if order_ws.cell(1, i).value is not None:
    #             graph_head.append(order_ws.cell(1, i).value)
    #         else:
    #             break
    #
    #     graph_folder = path + 'graph/'
    #     try:
    #         if not os.path.exists(graph_folder):
    #             os.makedirs(graph_folder)
    #             print("creat '%s' folder" % graph_folder)
    #     except OSError:
    #         print('Error: Creating directory. ' + graph_folder)
    #
    #
    #     for i in range(len(graph_head)):
    #         print(graph_head[i])
    #         graph_ws = summary_wb.create_sheet(graph_head[i])
    #         graph_list = []
    #
    #         for j in range(2, 1048576):
    #             if order_ws.cell(j, i + 1).value is not None:
    #                 temp = order_ws.cell(j, i + 1).value.lower()
    #                 if len(temp[3:]) == 1:
    #                     temp = 'tek' + '000' + temp[3:]
    #                     graph_list.append(temp)
    #                 elif len(temp[3:]) == 2:
    #                     temp = 'tek' + '00' + temp[3:]
    #                     graph_list.append(temp)
    #                 elif len(temp[3:]) == 3:
    #                     temp = 'tek' + '0' + temp[3:]
    #                     graph_list.append(temp)
    #                 elif len(temp[3:]) >= 5:
    #                     print('graph file name error!!!', end='    ')
    #                     print(order_ws.cell(j, i + 1).value)
    #                 else:
    #                     graph_list.append(temp)
    #             else:
    #                 break
    #
    #         absent_file = []
    #         for graph in graph_list:
    #             if graph + '.xlsx' not in file_list:
    #                 absent_file.append(graph)
    #                 graph_list.remove(graph)
    #         # graph_ws.cell(1, 1).value = absent_file
    #         for idx, file in enumerate(absent_file):
    #             graph_ws.cell(1, j + 2).value = file[idx]
    #
    #         for idx, graph in enumerate(graph_list):
    #             # wb = openpyxl.load_workbook(path + graph + '.xlsx')
    #             # ws = wb[graph]
    #             #
    #             # num_of_data_len = int(ws['b10'].value)
    #             # if num_of_data_len > 1000000 - 21:
    #             #     num_of_data_len = 1000000 - 21
    #             #
    #             # for j in range(2, 100000):
    #             #     if ws.cell(13, j).value is None:
    #             #         max_cal = j - 1
    #             #         break
    #             #
    #             # chart1 = openpyxl.chart.LineChart()
    #             # chart1.title = graph
    #             # chart1.style = 10
    #             # chart1.x_axis.title = "time"
    #             #
    #             # chart2 = openpyxl.chart.LineChart()
    #             # chart2.y_axis.majorGridlines = None
    #             # chart2.y_axis.axId = 200
    #             #
    #             # cats = openpyxl.chart.Reference(ws, min_col=1, min_row=22, max_row=num_of_data_len + 21)
    #             # for j in range(max_cal, max_cal-2, -1):
    #             #     if ws.cell(13, j).value == 'V':
    #             #         data1 = openpyxl.chart.Reference(ws, min_col=j, max_col=j, min_row=21,
    #             #                                          max_row=num_of_data_len + 21)
    #             #         print(type(data1))
    #             #         chart1.add_data(data1, titles_from_data=True)
    #             #         chart1.set_categories(cats)
    #             #     elif ws.cell(13, j).value == 'A':
    #             #         data2 = openpyxl.chart.Reference(ws, min_col=j, max_col=j, min_row=21,
    #             #                                          max_row=num_of_data_len + 21)
    #             #         print(help(openpyxl.chart.Reference))
    #             #         chart2.add_data(data2, titles_from_data=True)
    #             #         chart2.set_categories(cats)
    #             #
    #             #
    #             #
    #             # s1 = chart1.series[0]
    #             # s1.graphicalProperties.line.width = 0
    #             #
    #             # s2 = chart1.series[0]
    #             # s2.graphicalProperties.line.width = 0
    #             #
    #             # chart2.y_axis.crosses = "max"  # max인 축이 오른쪽에 위치
    #             #
    #             # chart1 += chart2
    #             #
    #             # chart_location_idx = 35 * idx
    #             #
    #             # chart_location = 'b' + str(4 + chart_location_idx)
    #             # graph_ws.add_chart(chart1, chart_location)
    #             # chart1.chart_width = 30
    #             # chart1.chart_height = 18
    #
    #             # wb.close()
    #
    #             excel = Dispatch('Excel.Application')
    #             excel.Visible = True
    #             wb = excel.Workbooks.Open(path + graph + '.xlsx')
    #             sheet = wb.Worksheets(graph)
    #             mychart = sheet.ChartObjects(1)
    #             mychart.Chart.Export(Filename=graph_folder + str(idx + 1) + ' - ' + graph + '.jpg')
    #         excel.Quit()

            # summary_wb.save(path + idx + ' - ' + summary + '.xlsx')

    def delete_by_df_column_value(self, new_sheet, data_items, df, sheet_name, filename, exclusive=True):
        if exclusive:
            for i in data_items[list(data_items.keys())[0]]:
                if list(data_items.keys())[0].lower() == 'pwm':
                    if 'PWM' in sheet_name:
                        sheet_name = sheet_name[:sheet_name.find('PWM')] + 'PWM_' + '{0:04d}'.format(int(i))
                    else:
                        sheet_name = sheet_name + ' ' + 'PWM_' + '{0:04d}'.format(int(i))
                elif list(data_items.keys())[0].lower() == 'ohm':
                    if 'ohm' in sheet_name:
                        sheet_name = sheet_name[:sheet_name.find('ohm') - 6] + ' ' + str(i) + 'ohm'
                    else:
                        sheet_name = sheet_name + ' ' + str(i) + 'ohm'
                elif list(data_items.keys())[0].lower() == 'ch':
                    if 'Ch' in sheet_name:
                        sheet_name = sheet_name[:sheet_name.find('Ch')] + str(i)
                    else:
                        sheet_name = sheet_name + ' ' + str(i)
                else:
                    if i in sheet_name:
                        sheet_name = sheet_name[:i]
                    else:
                        sheet_name = i


                sheet_name = sheet_name.lstrip()

                df_remain = df[df[list(data_items.keys())[0]] == i]

                if 0 != len(list(data_items.keys())) - 1:
                    data_items_delivery = data_items.copy()
                    del data_items_delivery[list(data_items_delivery.keys())[0]]
                    self.delete_by_df_column_value(new_sheet, data_items_delivery, df_remain, sheet_name, filename)
                else:
                    if not os.path.exists(self.tek_excel_path + filename): # excel.path로 변경
                        with pd.ExcelWriter(self.tek_excel_path + filename, mode='w', engine='openpyxl') as writer:
                                            df_remain.to_excel(writer, sheet_name='total', index=False)
                    else:
                        with pd.ExcelWriter(self.tek_excel_path + filename, mode='a', engine='openpyxl') as writer:
                            df_remain.to_excel(writer, sheet_name=sheet_name, index=False)
                    # df.to_excel(self.tek_excel_path + filename, sheet_name=sheet_name)
                    print(sheet_name)

        return df


    def get_seperated_data(self, file):
        items = pd.read_excel(self.path + self.eval_file, sheet_name='seperate summary')
        items = items.iloc[:, 0].tolist()
        for i in range(len(items) - 1, -1, -1):
            if type(items[i]) is float:
                del items[i]
            else:
                break

        df_summary = pd.read_excel(self.tek_excel_path + file, sheet_name='with kmon')
        data_items = {}

        for item in items:
            if item.lower() != 'and' and item.lower() != 'or':
                if item.lower() == 'pwm':
                    item = 'PWM'
                elif item.lower() == 'board':
                    item = 'Board'
                elif item.lower() == 'ohm':
                    item = 'ohm'
                elif item.lower() == 'ch':
                    item = 'Ch'

                label = list(df_summary[item].unique())
                if '-' in label:
                    label.remove('-')
                label.sort()
                data_items.setdefault(item, label)



        if items[-1].lower() == 'and':
            sheet_name = ''
            if 'Pwm' in items:
                items[items.index('Pwm')] = 'PWM'
            if items[0].lower() != 'board':
                sheet_clean = items[0].upper()
            else:
                sheet_clean = ''
            self.delete_by_df_column_value(sheet_clean, data_items, df_summary, sheet_name, file)


    def draw_chart(self, filename, sheet_head):
        wb = openpyxl.open(self.tek_excel_path + filename, read_only=False)
        sheets = wb.sheetnames
        sheets = [sheet for sheet in sheets if sheet[:2] == sheet_head[0]]

        data_length = 1
        for sheet in sheets:
            ws = wb[sheet]
            # i = 0
            while True:
                if ws.cell(data_length, 2).value is None:
                    break
                data_length += 1
            data_length -= 2

            for i in range(5):
                if ws.cell(data_length + 1, 11) is None:
                    ws.cell(data_length + 1, 11)
                if ws.cell(data_length + 1, 12) is None:
                    ws.cell(data_length + 1, 12)

            chart1 = openpyxl.chart.LineChart()
            chart2 = openpyxl.chart.LineChart()

            chart1.title = ws.cell(1, 11).value + '  -  ' + ws.cell(1, 12).value
            chart1.style = 10
            # chart1.series[0].marker.symbol = 'circle'
            # chart1.series[0].marker.graphicalProperties.solidFill = "FF0000"

            chart1.x_axis.title = ws.cell(1, 6).value
            # chart1.y_axis.scaling.min = v_min
            # chart1.y_axis.scaling.max = v_max
            data1 = openpyxl.chart.Reference(ws, min_col=11,
                                             min_row=1, max_row=data_length + 1)
            cats = openpyxl.chart.Reference(ws, min_col=6, min_row=2, max_row=data_length + 1)  # 축 설정
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.series[0].graphicalProperties.line.width = 0
            chart1.y_axis.majorGridlines = None

            chart2.style = 10
            chart2.x_axis.title = ws.cell(1, 11).value
            # chart1.y_axis.scaling.min = v_min
            # chart1.y_axis.scaling.max = v_max
            data2 = openpyxl.chart.Reference(ws, min_col=12,
                                             min_row=1, max_row=data_length + 1)
            cats = openpyxl.chart.Reference(ws, min_col=6, min_row=2, max_row=data_length + 1)  # 축 설정
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.series[0].graphicalProperties.line.width = 0
            chart2.y_axis.majorGridlines = None

            chart2.y_axis.axId = 200
            chart2.y_axis.crosses = 'max'

            chart1 += chart2

            chart_location = str(data_length + 5)
            ws.add_chart(chart1, 'c' + chart_location)
            chart1.width = 30
            chart1.height = 18

        wb.save(self.tek_excel_path + filename)


    def combine_kmon_data(self, file):
        df = 1
        num = -1
        info_test = {}
        previous = -1

        kmon_file = file[:5] + 'kmon' + file[9:]

        files = os.listdir(self.kmon_csv_path)
        files.sort()

        files = [file for file in files if file[-4:] == '.csv']

        for file in files:
            file = file.split('.csv')[0]
            if kmon_file in file:
                if info_test.get(kmon_file.split('_')[2]):
                    info_test[kmon_file.split('_')[2]].append(file.split('_')[3])
                else:
                    info_test.setdefault(kmon_file.split('_')[2], [file.split('_')[3]])

        df_num = []
        df = 0
        for idx, key in enumerate(info_test):
            previous = -1
            if previous != key:
                for i, value in enumerate(info_test[key]):
                    if i:
                        df_1 = pd.read_csv(self.kmon_csv_path + 'info_kmon_' + key + '_' + value + '.csv')
                        df = pd.concat([df, df_1], ignore_index=True, axis=1)
                    else:
                        df = pd.read_csv(self.kmon_csv_path + 'info_kmon_' + key + '_' + value + '.csv')
                df_num.append(df.copy())

        df = pd.DataFrame()
        for i in range(len(df_num[0].columns)):
            df[i] = np.nan

        # for i in range(len(df) - 1, -1, -1):
        #     df.drop([df.index[i]], inplace=True)

        for i in range(len(df_num)):
            df = pd.concat([df, df_num[i]], ignore_index=True)


        for i in range(len(df.columns) - 1, 0, -1):
            if i % 2 == 0:
                df.drop([df.columns[i]], axis=1, inplace=True)

        # # 왜 넣읐을까.. 삭제 대상
        # for i in range(len(df.columns) - 1, 0, -1):
        #     if df[df.columns[i]].max() == 0 and df[df.columns[i]].min():
        #         print('asdfasf')

        df_kmon_set = pd.read_excel(self.path + self.eval_file, sheet_name='kmon monitoring set RFamp')

        for i in range(len(df_kmon_set), 20):
            df_kmon_set = df_kmon_set.append(pd.Series(name=i))

        columns_name = []
        for idx, values in enumerate(df_kmon_set.columns):
            for idx, value in enumerate(df_kmon_set[values]):
                columns_name.append(value)

        for i in range(1, len(list(df.columns))):
            df = df.rename(columns={list(df.columns)[i]: columns_name[i - 1]})

        for i in range(len(df.columns) - 1, 1, -1):
            if type(df.columns[i]) is not str:
                df = df.drop(df.columns[i], axis=1)
                break

        df = df.drop(df.columns[0], axis=1)

        del df_1
        del df_kmon_set
        del df_num
        gc.collect()

        filename = kmon_file + '.xlsx'
        try:
            if not os.path.exists(self.kmon_csv_path + filename):
                with pd.ExcelWriter(self.kmon_csv_path + filename, mode='w', engine='openpyxl') as writer:
                    df.to_excel(writer)
            else:
                with pd.ExcelWriter(self.kmon_csv_path + filename, mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer)
        except:
            print('can not save kmon excel file')

        return df

    def combine_kmon_data_for_PL150(self, file):
        print('start combine_kmon_data_for_PL150')



    def check_kmon_and_testfile(self, df_kmon, test_file):
        test_info_files = pd.read_excel(self.path + self.eval_file, sheet_name='info_test files')
        test_info_files = test_info_files.iloc[:, 0].tolist()
        evaluation_set = pd.read_excel(self.path + self.eval_file, sheet_name='evaluation set')
        evaluation_set = evaluation_set.iloc[:, 0].tolist()

        absent_list = []
        control_value = {}
        control_value_pre = {}
        for control in evaluation_set:
            if control[:-1] != 'CP Pwm Ch ':
                control_value.setdefault(control)
                control_value_pre.setdefault(control)
            else:
                control_value.setdefault('CP Pwm Set Ch ' + control[-1])
                control_value_pre.setdefault('CP Pwm Set Ch ' + control[-1])

        add_item = ['deviation']     # # kmon 수집 항목에서 추가 계하여 넣을 항목들
        measure_value = {}        # # kmon 수집항목 중 필요한 것만 사용하기 위함
        measure_item = []
        for item in df_kmon.columns.tolist():
            if not (item in evaluation_set):
                measure_item.append(item)
                measure_value.setdefault(item, [])
                for add in add_item:
                    measure_value.setdefault(item + ' ' + add, [])
                self.measure_value.setdefault(item, [])
                for add in add_item:
                    self.measure_value.setdefault(item + ' ' + add, [])

        row = 0
        try:
            df_test_file = pd.read_excel(self.test_info_path + test_file + '.xlsx')
            print("open test list:\t{}".format(test_file))
        except:
            print("can't find test list {}.".format(test_file))

        # # # kmon 수집 항목 초기화
        # for key in measure_value.keys():
        #     measure_value[key] = []

        filenames = []

        for i in range(len(df_test_file)):
            for item in evaluation_set:
                try:
                    control_value[item] = df_test_file.at[i, item]
                except:
                    print(item)

                    print("테스트 파일에 있는 항목 {}과 제어내용이 동일하지 않습니다.".format(item))

            while True:
                for key, value in control_value.items():
                    if key != 'filename':
                        all_same = True
                        if df_kmon.at[row, key] != value:
                            all_same = False
                            if row != len(df_kmon) - 1:
                                row += 1
                                break
                            else:
                                break

                if not all_same and row == len(df_kmon) - 1:
                    break

                if all_same:
                    row += 1
                    same_test_condition = 1
                    while all_same and row != len(df_kmon):
                        for key, value in control_value.items():
                            if df_kmon.at[row, key] == value:
                                pass
                            else:
                                all_same = False
                                break
                        if all_same:
                            same_test_condition += 1
                            row += 1
                            if row == len(df_kmon):
                                break

                    for key in measure_item:
                        temp = []
                        for k in range(row - same_test_condition, row):
                            temp.append(df_kmon.at[k, key])
                        temp.pop(0)
                        standardization = np.std(temp)

                        for j in add_item:
                            measure_value[key + ' ' + j].append(standardization)

                        # if standardization > 5:
                        #     std = False
                        #     while not std:
                        #         odd = True
                        #         if odd:
                        #             temp.remove(min(temp))
                        #             odd = False
                        #         else:
                        #             temp.remove(max(temp))
                        #             odd = True
                        #         standardization = np.std(temp)
                        #         if standardization < 5:
                        #             std = True
                        try:
                            measure_value[key].append(np.mean(temp))
                        except:
                            measure_value[key].append(np.nan)
                            print("값들의 편차가 클 가능성이 높습니다. 확인 필요합니다. {}:\t{}".format(test_file, key))
                    filenames.append(df_test_file.at[i, df_test_file.columns[0]])
                    break

            if row == len(df_kmon) - 1:
                break

        self.measure_value['filename'].extend(filenames)

        for key, value in self.measure_value.items():
            if key != 'filename':
                self.measure_value[key].extend(measure_value[key])

        for i in range(len(df_test_file['filename'])):
            if not (df_test_file.at[i, 'filename'] in filenames):
                self.lost_files.append(df_test_file.at[i, 'filename'])

        if len(self.lost_files):
            print("{}에서 아래 파일일들 kmon 정보와 맞지 않습니다.".fomat(test_file))
            for i in range(len(self.lost_files)):
                print(self.lost_files[i])


    def merge_kmon_and_summary(self):
        try:
            df_summary = pd.read_excel(self.tek_excel_path + 'summary.xlsx', sheet_name='summary')
        except:
            print('something wrong: summary')

        for key in self.measure_value.keys():
            if key != 'filename':
                df_summary[key] = np.nan

        for i in range(len(self.measure_value['filename'])):
            for j in range(len(df_summary)):
                if self.measure_value['filename'][i] == df_summary.at[j, 'filename']:
                    for key in self.measure_value.keys():
                        df_summary.at[j, key] = self.measure_value[key][i]
                    break

        filename = 'summary.xlsx'
        if not os.path.exists(self.tek_excel_path + filename):
            with pd.ExcelWriter(self.tek_excel_path + filename, mode='w', engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='with kmon', index=False)
        else:
            with pd.ExcelWriter(self.tek_excel_path + filename, mode='a', engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='with kmon', index=False)


    def kmon_change_digit(self, sheet_name):
        df_ctrl = pd.read_excel(self.path + self.eval_file, sheet_name=sheet_name)
        kmon_files = df_ctrl.iloc[:, 0]
        kmon_files = kmon_files.dropna()
        df_ctrl.drop([df_ctrl.columns[0]], axis=1, inplace=True)

        for file in kmon_files:
            file = os.path.join(self.kmon_csv_path, file + '.xlsx')
            df_kmon = pd.read_excel(file)
            df_kmon.drop([df_kmon.columns[0]], axis=1, inplace=True)

            for i in range(len(df_ctrl)):
                name, digit = df_ctrl.iloc[i, :].Name, df_ctrl.iloc[i, :].digit
                command, digit = digit.split(' ')[0], int(digit.split(' ')[1])
                for row, data in enumerate(df_kmon[name]):
                    if command.lower() == 'reduce':
                        df_kmon.at[row, name] = data / (digit * 10)
                    elif (command.lower() == 'increase') or (command.lower() =='raise') or (command.lower() == 'expand'):
                        df_kmon.at[row, name] = data * (digit * 10)

            try:
                if not os.path.exists(file):
                    with pd.ExcelWriter(file, mode='w', engine='openpyxl') as writer:
                        df_kmon.to_excel(writer, sheet_name='changed digit', index=True)
                else:
                    with pd.ExcelWriter(file, mode='a', engine='openpyxl') as writer:
                        df_kmon.to_excel(writer, sheet_name='changed digit', index=True)
            except:
                print('can not save kmon excel file:', file)




if __name__=='__main__':
    if platform.platform()[:3].lower() == 'mac':
        mac_m1 = True
    elif platform.platform()[:3].lower() == 'win':
        mac_m1 = False

    if mac_m1:
        path = '/Users/rainyseason/winston/Workspace/python/Pycharm Project/sinewave_analyze/Evaluation/'
        path_csv = path + 'tek_csv/'
        path_excel = path + 'tek_excel/'
        path_summary = path + 'summary/'
        path_information = path + 'test information/'
        path_kmon = path + 'kmon_csv'
    else:
        path = 'D:/data_analyze/'
        path_csv = path + 'tek_csv/'
        path_excel = path + 'tek_excel/'
        path_summary = path + 'summary/'
        path_information = path + 'test information/'
        path_kmon = path + 'kmon_csv/'

    evaluation_control_file = 'eval_control.xlsx'

    merge_kmon = Get_Summary(path, evaluation_control_file)
    merge_kmon.kmon_change_digit('kmon digit change set RFamp')
